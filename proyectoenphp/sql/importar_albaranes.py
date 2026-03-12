"""
importar_albaranes.py
=====================
Importa albaranes de venta y compra desde Excel a la BD MySQL (licitaciones).

Tablas destino:
  - tbl_albaranes        → cabecera del albarán
  - tbl_albaranes_lineas → líneas con precio_unitario (PVU en ventas, PCU en compras)

Tras importar, las vistas v_ultimo_pvu / v_ultimo_pcu / v_margen_producto
quedan automáticamente actualizadas.

Uso:
  pip install openpyxl pymysql
  python importar_albaranes.py

Configura la sección DATABASE antes de ejecutar.
"""

import sys
import logging
from datetime import datetime, date
from pathlib import Path

import openpyxl
import pymysql
import pymysql.cursors

# ─────────────────────────────────────────────
#  CONFIGURACIÓN — edita esto antes de ejecutar
# ─────────────────────────────────────────────
DATABASE = {
    "host":     "127.0.0.1",
    "port":     3308,
    "user":     "root",
    "password": "",
    "db":       "licitaciones",
    "charset":  "utf8mb4",
}

# Rutas a los ficheros Excel (absolutas o relativas al script)
EXCEL_VENTA  = Path("Total_albaranes_venta.xlsm")
EXCEL_COMPRA = Path("Albaranes_de_compra.xlsx")

# Si True, solo muestra qué haría sin tocar la BD
DRY_RUN = False

# Tamaño de lote para commits parciales (reduce memoria en transacciones grandes)
BATCH_SIZE = 500
# ─────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ══════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════

def to_date(val) -> date | None:
    """Convierte datetime de openpyxl a date de Python."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def to_str(val, max_len: int = None) -> str | None:
    """Convierte a str, recorta si hace falta."""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    if max_len and len(s) > max_len:
        s = s[:max_len]
    return s


def to_decimal(val) -> float | None:
    """Convierte a float seguro."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def lote_str(val) -> str | None:
    """Los lotes vienen como int en Excel; los guardamos como string."""
    if val is None:
        return None
    return str(int(val)) if isinstance(val, (int, float)) else to_str(val, 100)


# ══════════════════════════════════════════════
#  CACHÉ DE IDs
# ══════════════════════════════════════════════

class Cache:
    """
    Mantiene en memoria los ids ya conocidos para evitar SELECT repetidos.
    """
    def __init__(self, conn):
        self.conn = conn
        self._productos: dict[str, int] = {}   # referencia → id
        self._contactos: dict[str, int] = {}   # nombre → id
        self._albaranes: dict[tuple, int] = {}  # (numero, tipo) → id_albaran
        self._refs_sin_producto: set[str] = set()

    # ── Productos ──────────────────────────────
    def get_id_producto(self, ref: str) -> int | None:
        if ref in self._refs_sin_producto:
            return None
        if ref not in self._productos:
            with self.conn.cursor() as cur:
                cur.execute(
                    "SELECT id FROM tbl_productos WHERE referencia = %s LIMIT 1",
                    (ref,)
                )
                row = cur.fetchone()
            if row:
                self._productos[ref] = row["id"]
            else:
                self._refs_sin_producto.add(ref)
                return None
        return self._productos[ref]

    # ── Contactos ──────────────────────────────
    def get_id_contacto(self, nombre: str) -> int | None:
        """
        Busca el contacto por nombre (campo `nombre` en contactos).
        Los nombres del Excel suelen venir como 'NOMBRE FISCAL / NOMBRE COMERCIAL';
        probamos primero el nombre completo y luego solo la parte antes de ' / '.
        """
        if not nombre:
            return None
        if nombre in self._contactos:
            return self._contactos[nombre]

        candidatos = [nombre]
        if " / " in nombre:
            candidatos.append(nombre.split(" / ")[0].strip())

        with self.conn.cursor() as cur:
            for candidato in candidatos:
                cur.execute(
                    "SELECT id FROM contactos WHERE nombre = %s OR nombre_fiscal = %s LIMIT 1",
                    (candidato, candidato)
                )
                row = cur.fetchone()
                if row:
                    self._contactos[nombre] = row["id"]
                    return row["id"]

        # No encontrado — lo dejamos como NULL (guardamos None para no repetir búsquedas)
        self._contactos[nombre] = None
        return None

    # ── Albaranes (cabecera) ───────────────────
    def albaran_exists(self, numero: str, tipo: str) -> int | None:
        key = (numero, tipo)
        if key not in self._albaranes:
            with self.conn.cursor() as cur:
                cur.execute(
                    "SELECT id_albaran FROM tbl_albaranes "
                    "WHERE numero_albaran = %s AND tipo_albaran = %s LIMIT 1",
                    (numero, tipo)
                )
                row = cur.fetchone()
            self._albaranes[key] = row["id_albaran"] if row else None
        return self._albaranes[key]

    def set_albaran(self, numero: str, tipo: str, id_albaran: int):
        self._albaranes[(numero, tipo)] = id_albaran


# ══════════════════════════════════════════════
#  IMPORTACIÓN
# ══════════════════════════════════════════════

def insert_albaran(cur, cache: Cache, numero: str, tipo: str, fecha: date,
                   num_factura: str | None, nombre_contacto: str | None,
                   comercial: str | None) -> int:
    """
    Inserta la cabecera del albarán si no existe.
    Devuelve el id_albaran.
    """
    existing = cache.albaran_exists(numero, tipo)
    if existing:
        return existing

    id_contacto = cache.get_id_contacto(nombre_contacto) if nombre_contacto else None

    cur.execute(
        """
        INSERT INTO tbl_albaranes
            (numero_albaran, tipo_albaran, fecha_albaran,
             numero_factura, id_contacto, nombre_contacto, comercial)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        """,
        (numero, tipo, fecha, num_factura, id_contacto,
         to_str(nombre_contacto, 255), to_str(comercial, 255))
    )
    id_albaran = cur.lastrowid
    cache.set_albaran(numero, tipo, id_albaran)
    return id_albaran


def insert_linea(cur, id_albaran: int, id_producto: int | None,
                 ref: str, nombre_art: str, familia: str | None,
                 lote: str | None, cantidad: float, precio: float | None,
                 iva_pct: float | None, dto_pct: float | None, importe: float | None):
    cur.execute(
        """
        INSERT INTO tbl_albaranes_lineas
            (id_albaran, id_producto, ref_articulo, nombre_articulo,
             familia, lote, cantidad, precio_unitario,
             iva_pct, descuento_pct, importe)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """,
        (id_albaran, id_producto,
         to_str(ref, 255), to_str(nombre_art, 500),
         to_str(familia, 255), lote,
         cantidad, precio,
         iva_pct, dto_pct, importe)
    )


def importar_venta(conn, cache: Cache):
    """
    Importa Total_albaranes_venta.xlsm
    Columnas: Nº Albarán · Fecha · Nº Factura · Cliente · Comercial ·
              Familia · Ref. Artículo · Lote · Artículo · Cantidad ·
              Precio (→ PVU) · %IVA · % Dto. · Importe
    """
    log.info("Abriendo %s …", EXCEL_VENTA)
    wb = openpyxl.load_workbook(EXCEL_VENTA, read_only=True, data_only=True)
    ws = wb.active

    stats = {"filas": 0, "lineas_ok": 0, "lineas_sin_ref": 0,
             "albaranes_nuevos": 0, "errores": 0}

    with conn.cursor() as cur:
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Desempaquetar columnas
            num_alb, fecha, num_fac, cliente, comercial, \
            familia, ref, lote, articulo, cantidad, precio, \
            iva, dto, importe = row

            # Saltar filas vacías
            if not num_alb or not fecha:
                continue

            stats["filas"] += 1
            fecha_d = to_date(fecha)
            ref      = to_str(ref, 255)
            cantidad = to_decimal(cantidad) or 0.0

            try:
                # Cabecera
                id_albaran = insert_albaran(
                    cur, cache,
                    numero=to_str(num_alb, 50),
                    tipo="VENTA",
                    fecha=fecha_d,
                    num_factura=to_str(num_fac, 50),
                    nombre_contacto=to_str(cliente, 255),
                    comercial=to_str(comercial, 255),
                )

                # Línea
                id_prod = cache.get_id_producto(ref) if ref else None
                if not ref or id_prod is None:
                    stats["lineas_sin_ref"] += 1

                insert_linea(
                    cur, id_albaran, id_prod,
                    ref=ref or "",
                    nombre_art=to_str(articulo, 500) or "",
                    familia=to_str(familia, 255),
                    lote=lote_str(lote),
                    cantidad=cantidad,
                    precio=to_decimal(precio),   # PVU
                    iva_pct=to_decimal(iva),
                    dto_pct=to_decimal(dto),
                    importe=to_decimal(importe),
                )
                stats["lineas_ok"] += 1

            except Exception as exc:
                stats["errores"] += 1
                log.warning("Fila %d (VENTA) error: %s | datos: %s", i, exc, row)
                conn.rollback()
                continue

            # Commit por lotes
            if stats["filas"] % BATCH_SIZE == 0:
                if not DRY_RUN:
                    conn.commit()
                log.info("  VENTA: %d filas procesadas …", stats["filas"])

    if not DRY_RUN:
        conn.commit()
    wb.close()
    log.info("VENTA completado: %s", stats)
    return stats


def importar_compra(conn, cache: Cache):
    """
    Importa Albaranes_de_compra.xlsx
    Columnas: Fecha · Nº Albarán · Proveedor · Familia · Ref. Artículo ·
              Artículo · Cantidad · Precio · Precio GI (→ PCU real) ·
              %IVA · % Dto. · Importe

    Precio GI es el precio tras aplicar descuento global del proveedor,
    que es el coste unitario real → se guarda en precio_unitario.
    """
    log.info("Abriendo %s …", EXCEL_COMPRA)
    wb = openpyxl.load_workbook(EXCEL_COMPRA, read_only=True, data_only=True)
    ws = wb.active

    stats = {"filas": 0, "lineas_ok": 0, "lineas_sin_ref": 0,
             "albaranes_nuevos": 0, "errores": 0}

    with conn.cursor() as cur:
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            fecha, num_alb, proveedor, familia, ref, articulo, \
            cantidad, precio_bruto, precio_gi, iva, dto, importe = row

            if not num_alb or not fecha:
                continue

            stats["filas"] += 1
            fecha_d  = to_date(fecha)
            ref      = to_str(ref, 255)
            cantidad = to_decimal(cantidad) or 0.0

            # PCU = Precio GI (precio neto tras descuento de proveedor)
            # Si Precio GI es None, caemos al precio bruto
            pcu = to_decimal(precio_gi) or to_decimal(precio_bruto)

            try:
                id_albaran = insert_albaran(
                    cur, cache,
                    numero=to_str(num_alb, 50),
                    tipo="COMPRA",
                    fecha=fecha_d,
                    num_factura=None,          # las compras no llevan nº factura en el Excel
                    nombre_contacto=to_str(proveedor, 255),
                    comercial=None,
                )

                id_prod = cache.get_id_producto(ref) if ref else None
                if not ref or id_prod is None:
                    stats["lineas_sin_ref"] += 1

                insert_linea(
                    cur, id_albaran, id_prod,
                    ref=ref or "",
                    nombre_art=to_str(articulo, 500) or "",
                    familia=to_str(familia, 255),
                    lote=None,          # compras no tienen lote en el Excel
                    cantidad=cantidad,
                    precio=pcu,         # PCU
                    iva_pct=to_decimal(iva),
                    dto_pct=to_decimal(dto),
                    importe=to_decimal(importe),
                )
                stats["lineas_ok"] += 1

            except Exception as exc:
                stats["errores"] += 1
                log.warning("Fila %d (COMPRA) error: %s | datos: %s", i, exc, row)
                conn.rollback()
                continue

            if stats["filas"] % BATCH_SIZE == 0:
                if not DRY_RUN:
                    conn.commit()
                log.info("  COMPRA: %d filas procesadas …", stats["filas"])

    if not DRY_RUN:
        conn.commit()
    wb.close()
    log.info("COMPRA completado: %s", stats)
    return stats


# ══════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════

def main():
    # Validar que los ficheros existen
    for f in (EXCEL_VENTA, EXCEL_COMPRA):
        if not f.exists():
            log.error("No se encuentra el fichero: %s", f.resolve())
            sys.exit(1)

    if DRY_RUN:
        log.warning("═══ MODO DRY RUN — no se escribirá nada en la BD ═══")

    log.info("Conectando a MySQL en %s:%s/%s …", DATABASE["host"], DATABASE["port"], DATABASE["db"])
    conn = pymysql.connect(
        **DATABASE,
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=False,
    )
    log.info("Conexión OK.")

    cache = Cache(conn)

    try:
        t0 = datetime.now()

        stats_v = importar_venta(conn, cache)
        stats_c = importar_compra(conn, cache)

        elapsed = (datetime.now() - t0).total_seconds()

        log.info("═══════════════════════════════════════════")
        log.info("RESUMEN FINAL (%.1f s)", elapsed)
        log.info("  VENTA  → %d líneas importadas, %d sin id_producto, %d errores",
                 stats_v["lineas_ok"], stats_v["lineas_sin_ref"], stats_v["errores"])
        log.info("  COMPRA → %d líneas importadas, %d sin id_producto, %d errores",
                 stats_c["lineas_ok"], stats_c["lineas_sin_ref"], stats_c["errores"])
        log.info("  Referencias sin producto en tbl_productos: %d únicas",
                 len(cache._refs_sin_producto))
        if cache._refs_sin_producto:
            muestra = list(cache._refs_sin_producto)[:10]
            log.info("    Muestra: %s", muestra)
        log.info("═══════════════════════════════════════════")
        log.info("Las vistas v_ultimo_pcu, v_ultimo_pvu y v_margen_producto")
        log.info("ya reflejan los nuevos datos. Puedes consultarlas directamente.")

    finally:
        conn.close()


if __name__ == "__main__":
    main()